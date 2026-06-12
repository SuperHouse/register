import datetime
import re
import string

from django.core.management.base import BaseCommand
from django.utils import timezone
from openpyxl import load_workbook

from device.models import (
    Design,
    Device,
    DeviceEvent,
    TestRecord,
)
from crm.models import Org
from utils import date_from_str, int_map

# timezone.activate(tz)


class Command(BaseCommand):
    help = 'Import data from an XLSX file'
    output_transaction = True  # Put all db ops in a transaction

    def add_arguments(self, parser):
        # parser.add_argument('filename', nargs='?', help='XLSX file to import from')
        parser.add_argument('filename', help='XLSX file to import from')
        # parser.add_argument('dryrun', type=bool, help='Dry run (no db changes)')
        pass

    def handle(self, *args, **options):
        filename = options['filename']
        # filename = 'pyproj/stash/Device Serial Numbers-3.xlsx'
        self.stdout.write(f'Importing from {filename}')

        wb = load_workbook(filename=filename)

        known_sheets = 'Devices/Queue/DeviceTypes/Raw Serials/Patched Boards'
        assert set(known_sheets.split('/')) == set(
            wb.sheetnames
        ), f'Expected sheets: {known_sheets}; Actual sheets: {wb.sheetnames}'

        # Read from the second sheet, DeviceTypes
        ws_design = wb['DeviceTypes']

        known_design_keys = tuple('Serial/ClientSerial/SKU/Name/HW Version/Customer/Price/Price2'.split('/'))
        design_keys = tuple(cell.value for cell in ws_design['1'] if cell.value is not None)
        # self.stdout.write(f'{design_keys=}')
        assert design_keys == known_design_keys

        col_map = {key: letter for letter, key in zip(string.ascii_uppercase, known_design_keys)}
        # self.stdout.write(f'{col_map=}')

        client_serials = tuple(int_map(cell.value) for cell in ws_design[col_map["ClientSerial"]][1:])
        client_names = tuple(cell.value for cell in ws_design[col_map['Customer']][1:])
        client_info = zip(client_serials, client_names)
        client_map = {}
        for row_nr, (client_id, name) in enumerate(client_info):
            row_nr += 1
            if name is None:
                continue
            # self.stdout.write(f'{row_nr=} {client_id=} {name=}')
            if client_id in client_map:
                assert client_map[client_id] == name, f"For serial {row_nr}, id {client_id} with name {name} doesn't match id {client_id} with name {client_map[client_id]} in database"
            else:
                client_map[client_id] = name

        # self.stdout.write(f'{client_map}')

        # No longer deleting any records - we'll update existing ones and add new ones

        # Import / update clients
        for id, name in client_map.items():
            try:
                client = Org.objects.get(pk=id)
                # print(f'Name in SQL: {client.company_name}; Name in XLSX: {name}')
                # If we wanted to update client records in the db from the sheet, do it here.
            except Org.DoesNotExist:
                # New client
                client = Org(pk=id, company_name=name)
                client.save()

        # Import designs
        design_count = 0
        design_updated_count = 0
        for row in ws_design.iter_rows(min_row=2, max_col=len(col_map.keys()), max_row=999):
            row = dict(zip(design_keys, (cell.value for cell in row)))
            # self.stdout.write(f'{row=}')
            if row['SKU'] is None:
                break
            id = int(row['Serial'])
            client_id = int(row['ClientSerial']) if row['ClientSerial'] else next_client_id
            design_data = {
                'id': id,
                'client_id': client_id,
                'sku': row['SKU'],
                'name': row['Name'],
                'hw_version': row['HW Version'],
            }
            if row['Price']:
                design_data['price'] = row['Price']
            if row['Price2']:
                design_data['price2'] = row['Price2']
            
            # Check if design exists, update if it does, create if it doesn't
            try:
                design = Design.objects.get(pk=id)
                # Update existing design
                design.client_id = design_data['client_id']
                design.sku = design_data['sku']
                design.name = design_data['name']
                design.hw_version = design_data['hw_version']
                if 'price' in design_data:
                    design.price = design_data['price']
                if 'price2' in design_data:
                    design.price2 = design_data['price2']
                design.save()
                design_updated_count += 1
            except Design.DoesNotExist:
                # Create new design
                design = Design(**design_data)
                design.save()
                design_count += 1

        # Now read from the first sheet, Devices
        ws_device = wb['Devices']

        known_device_keys = tuple(
            'Serial/DeviceTypeSerial/Assembled/Tested/Firmware/Notes/Device/HW Version/Invoice/PO/Shipped'.split('/')
        )
        ignore_device_keys = tuple(
            'ExtAddr/Truck/Location/Compound'.split('/')
        )
        self.stdout.write(self.style.WARNING(f'Ignoring these columns: {ignore_device_keys}'))
        device_keys = tuple(cell.value for cell in ws_device['1'] if cell.value is not None and cell.value not in ignore_device_keys)
        # self.stdout.write(f'{device_keys=}')
        assert device_keys == known_device_keys, f"Unexpected column(s): {device_keys=} {known_device_keys=}"

        col_map = {key: letter for letter, key in zip(string.ascii_uppercase, known_device_keys)}
        # self.stdout.write(f'{col_map=}')

        # Import devices
        tr = []
        de_list = []
        tr_list = []
        device_count = 0
        device_updated_count = 0
        for row in ws_device.iter_rows(min_row=2, max_col=len(col_map.keys()), max_row=9999):
            row = dict(zip(device_keys, (cell.value for cell in row)))
            if row['DeviceTypeSerial'] is None:
                continue
            # self.stdout.write(f'{row=}')
            device_id = int(row['Serial'])
            design_id = int(row['DeviceTypeSerial'])
            tested = row['Tested']
            sw_version = row['Firmware']
            creation_dt = row['Assembled']
            notes = row['Notes']
            invoice = row['Invoice']
            shipping = row['Shipped']
            porder = row['PO']

            if invoice:
                try:
                    # If the invoice is a straight number, it'll have .0 on the end.  Convert to string via int.
                    invoice = str(int(invoice))
                except ValueError:
                    # Just take the string, may not have been a straight number.
                    invoice = str(invoice)

            if porder:
                try:
                    # If the porder is a straight number, it'll have .0 on the end.  Convert to string via int.
                    porder = str(int(porder))
                except ValueError:
                    # Just take the string, may not have been a straight number.
                    porder = str(porder)

            if notes:
                dated_event_in_note_matchers = (
                    r'(\d{1,2}-[A-Z][a-z]{2}-202\d)',
                    r'(\d+/\d+/202\d)',
                )
                for matcher in dated_event_in_note_matchers:
                    match = re.search(matcher, notes)
                    if match:
                        # self.stdout.write(f'{row=}')
                        # self.stdout.write(f'{notes=}')
                        # d = match.groupdict()
                        # d_str = f'{d=}'
                        # self.stdout.write(d_str)
                        de_data = {
                            'device_id': device_id,
                            'event_dt': date_from_str(match.group()),
                            'event_type': 'NOTE',
                            'description': notes,
                        }
                        de_list.append(de_data)
                        notes = None
                        break

            if shipping:
                match = re.search(r'(\d{1,2}-[A-Z][a-z]{2}-20[23]\d):?( (.*))?', shipping)
                assert match, f"Oops, failed to match shipping.  {shipping=} {row=}"
                # self.stdout.write(f'{row=}')
                # self.stdout.write(f'{shipping=}')
                # d = match.groupdict()
                # d_str = f'{d=}'
                # self.stdout.write(d_str)
                de_data = {
                    'device_id': device_id,
                    'event_dt': date_from_str(match.group(1)),
                    'event_type': 'SHIPPING',
                    'description': match.group(3),
                }
                de_list.append(de_data)

            if tested:
                assert type(tested) is datetime.datetime
                test_dt = timezone.make_aware(datetime.datetime.combine(tested.date(), datetime.time()))
                tr_data = {
                    'device_id': device_id,
                    'test_dt': test_dt,
                    'result': TestRecord.PASS,
                    'notes': 'Test date imported from spreadsheet.',
                }
                tr_list.append(tr_data)

            device_data = {
                'id': device_id,
                'design_id': design_id,
                'creation_dt': timezone.make_aware(datetime.datetime.combine(creation_dt.date(), datetime.time())),
                'invoice': invoice,
                'po': porder,
                'notes': notes,
            }

            # Check if device exists, update if it does, create if it doesn't
            try:
                device = Device.objects.get(pk=device_id)
                # Update existing device
                device.design_id = device_data['design_id']
                device.creation_dt = device_data['creation_dt']
                device.invoice = device_data['invoice']
                device.po = device_data['po']
                device.notes = device_data['notes']
                device.save()
                device_updated_count += 1
            except Device.DoesNotExist:
                # Create new device
                device = Device(**device_data)
                device.save()
                device_count += 1

        # Import device events (avoid duplicates)
        de_created_count = 0
        de_skipped_count = 0
        for de_data in de_list:
            # Check if event already exists (same device, event_type, description, and event_dt)
            existing = DeviceEvent.objects.filter(
                device_id=de_data['device_id'],
                event_type=de_data['event_type'],
                description=de_data['description'],
                event_dt=de_data['event_dt']
            ).first()
            if not existing:
                de = DeviceEvent(**de_data)
                de.save()
                de_created_count += 1
            else:
                de_skipped_count += 1

        # Import test records (avoid duplicates)
        tr_created_count = 0
        tr_skipped_count = 0
        for tr_data in tr_list:
            # Check if test record already exists (same device and test_dt)
            existing = TestRecord.objects.filter(
                device_id=tr_data['device_id'],
                test_dt=tr_data['test_dt']
            ).first()
            if not existing:
                tr = TestRecord(**tr_data)
                tr.save()
                tr_created_count += 1
            else:
                tr_skipped_count += 1

        self.stdout.write(
            f'Imported {design_count} new designs, updated {design_updated_count} existing designs, '
            f'imported {device_count} new devices, updated {device_updated_count} existing devices, '
            f'created {de_created_count} device events (skipped {de_skipped_count} duplicates), '
            f'and created {tr_created_count} test records (skipped {tr_skipped_count} duplicates).'
        )

        self.stdout.write(self.style.SUCCESS('Done.'))
